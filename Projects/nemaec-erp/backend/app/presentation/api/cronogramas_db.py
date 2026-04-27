"""
📊 CRONOGRAMAS API ROUTER - DATABASE VERSION - NEMAEC ERP
Drop-in replacement for JSON-based cronogramas API using SQLAlchemy database.
Maintains exact same Pydantic models and response format as the JSON version.
"""
import io
from typing import List, Optional, Dict, Any
from datetime import datetime
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from pydantic import BaseModel
import openpyxl

from app.core.database import get_db
from app.infrastructure.database.models import CronogramaModel, PartidaModel

router = APIRouter(
    prefix="/cronogramas",
    tags=["cronogramas"],
    responses={404: {"description": "Not found"}}
)

# Modelos Pydantic - Idénticos a la versión JSON

class PartidaBase(BaseModel):
    codigo_interno: str
    comisaria_id: int
    codigo_partida: str
    descripcion: str
    unidad: str
    metrado: float
    precio_unitario: float
    precio_total: float
    fecha_inicio: Optional[str] = None  # 🎯 PERMITIR FECHAS NULAS
    fecha_fin: Optional[str] = None     # 🎯 PERMITIR FECHAS NULAS
    nivel_jerarquia: int
    partida_padre: Optional[str] = None

class PartidaResponse(PartidaBase):
    id: int
    cronograma_id: int
    created_at: str

class CronogramaCreate(BaseModel):
    comisaria_id: int
    nombre_cronograma: Optional[str] = None

class PartidaFechasUpdate(BaseModel):
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None

class CronogramaResponse(BaseModel):
    id: int
    comisaria_id: int
    nombre: str
    descripcion: Optional[str] = None
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    estado: str
    progreso: float
    created_at: str
    updated_at: Optional[str] = None

class CronogramaWithPartidas(CronogramaResponse):
    partidas: List[PartidaResponse]

class ImportStats(BaseModel):
    total_rows: int
    valid_partidas: int
    invalid_rows: int
    cronograma_id: int
    stats: Dict[str, Any]

# Helper functions

def partida_model_to_response(db_model: PartidaModel) -> PartidaResponse:
    """Convert SQLAlchemy PartidaModel to Pydantic response"""
    return PartidaResponse(
        id=db_model.id,
        cronograma_id=db_model.cronograma_id,
        codigo_interno=db_model.codigo_interno,
        comisaria_id=db_model.comisaria_id,
        codigo_partida=db_model.codigo_partida,
        descripcion=db_model.descripcion,
        unidad=db_model.unidad,
        metrado=db_model.metrado,
        precio_unitario=db_model.precio_unitario,
        precio_total=db_model.precio_total,
        fecha_inicio=db_model.fecha_inicio.isoformat() if db_model.fecha_inicio else None,
        fecha_fin=db_model.fecha_fin.isoformat() if db_model.fecha_fin else None,
        nivel_jerarquia=db_model.nivel_jerarquia,
        partida_padre=db_model.partida_padre,
        created_at=db_model.created_at.isoformat() if db_model.created_at else datetime.now().isoformat()
    )

def cronograma_model_to_response(db_model: CronogramaModel) -> CronogramaResponse:
    """Convert SQLAlchemy CronogramaModel to Pydantic response"""
    return CronogramaResponse(
        id=db_model.id,
        comisaria_id=db_model.comisaria_id,
        nombre=db_model.nombre,
        descripcion=db_model.descripcion,
        fecha_inicio=db_model.fecha_inicio.isoformat() if db_model.fecha_inicio else None,
        fecha_fin=db_model.fecha_fin.isoformat() if db_model.fecha_fin else None,
        estado=db_model.estado,
        progreso=db_model.progreso,
        created_at=db_model.created_at.isoformat() if db_model.created_at else datetime.now().isoformat(),
        updated_at=db_model.updated_at.isoformat() if db_model.updated_at else None
    )

def get_partida_padre(codigo_partida: str) -> Optional[str]:
    """Obtener código de partida padre"""
    partes = codigo_partida.split('.')
    if len(partes) > 1:
        return '.'.join(partes[:-1])
    return None

def calcular_nivel_jerarquia(codigo_partida: str) -> int:
    """Calcular nivel de jerarquía basado en el código de partida"""
    return len(codigo_partida.split('.'))

# Endpoints

@router.get("/", response_model=List[CronogramaResponse])
async def get_all_cronogramas(db: AsyncSession = Depends(get_db)):
    """
    Obtener todos los cronogramas

    Returns:
        List[CronogramaResponse]: Lista de cronogramas
    """
    print("🔍 DEBUG: GET /cronogramas - Consultando base de datos SQLite")

    stmt = select(CronogramaModel).order_by(CronogramaModel.created_at.desc())
    result = await db.execute(stmt)
    cronogramas = result.scalars().all()

    print(f"🗄️ Encontrados {len(cronogramas)} cronogramas en la base de datos")

    return [cronograma_model_to_response(cronograma) for cronograma in cronogramas]

@router.get("/comisaria/{comisaria_id}", response_model=List[CronogramaResponse])
async def get_cronogramas_by_comisaria(comisaria_id: int, db: AsyncSession = Depends(get_db)):
    """
    Obtener cronogramas de una comisaría específica

    Args:
        comisaria_id: ID de la comisaría

    Returns:
        List[CronogramaResponse]: Cronogramas de la comisaría
    """
    print(f"🔍 DEBUG: GET /cronogramas/comisaria/{comisaria_id}")

    stmt = select(CronogramaModel).where(
        CronogramaModel.comisaria_id == comisaria_id
    ).order_by(CronogramaModel.created_at.desc())

    result = await db.execute(stmt)
    cronogramas = result.scalars().all()

    print(f"🗄️ Encontrados {len(cronogramas)} cronogramas para comisaría {comisaria_id}")
    return [cronograma_model_to_response(cronograma) for cronograma in cronogramas]

@router.get("/{cronograma_id}", response_model=CronogramaWithPartidas)
async def get_cronograma_with_partidas(cronograma_id: int, db: AsyncSession = Depends(get_db)):
    """
    Obtener cronograma con sus partidas

    Args:
        cronograma_id: ID del cronograma

    Returns:
        CronogramaWithPartidas: Cronograma con lista de partidas
    """
    # Obtener cronograma
    stmt = select(CronogramaModel).where(CronogramaModel.id == cronograma_id)
    result = await db.execute(stmt)
    cronograma = result.scalar_one_or_none()

    if not cronograma:
        raise HTTPException(status_code=404, detail="Cronograma no encontrado")

    # Obtener partidas del cronograma
    stmt = select(PartidaModel).where(
        PartidaModel.cronograma_id == cronograma_id
    ).order_by(PartidaModel.codigo_partida)

    result = await db.execute(stmt)
    partidas = result.scalars().all()

    # Convertir a response format
    cronograma_response = cronograma_model_to_response(cronograma)
    partidas_response = [partida_model_to_response(partida) for partida in partidas]

    return CronogramaWithPartidas(
        **cronograma_response.model_dump(),
        partidas=partidas_response
    )

@router.post("/import", response_model=ImportStats, status_code=201)
async def import_cronograma_from_excel(
    comisaria_id: int = Form(...),
    nombre_cronograma: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Importar cronograma desde archivo Excel (formato SGP .xls o formatos anteriores .xlsx)

    Formato SGP (columnas A-K):
      A=comisaria, B=proveedor, C=codigo, D=partida, E=und,
      F=metrado, G=pu, H=parcial, I=met_eje, J=inicio, K=fec_termino

    Returns:
        ImportStats: Estadísticas de la importación
    """
    print(f"📊 Importando cronograma: {nombre_cronograma} para comisaría {comisaria_id}")

    filename_lower = file.filename.lower() if file.filename else ""
    print(f"DEBUG filename={file.filename!r} lower={filename_lower!r}")
    import sys; sys.stdout.flush()
    if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")

    is_old_xls = filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx')
    print(f"DEBUG is_old_xls={is_old_xls}")
    sys.stdout.flush()

    try:
        contents = await file.read()

        # Crear cronograma
        nuevo_cronograma = CronogramaModel(
            comisaria_id=comisaria_id,
            nombre=nombre_cronograma,
            descripcion=f"Cronograma importado desde {file.filename}",
            estado="activo",
            progreso=0.0,
            created_at=datetime.now()
        )
        db.add(nuevo_cronograma)
        await db.commit()
        await db.refresh(nuevo_cronograma)
        cronograma_id = nuevo_cronograma.id

        # ── Helpers ────────────────────────────────────────────────────────────

        def normalizar_codigo(val) -> str:
            import math
            if val is None:
                return ""
            if isinstance(val, str):
                s = val.strip()
                if s in ("nan", "None", ""):
                    return ""
                partes = s.split(".")
                result = []
                for i, p in enumerate(partes):
                    t = p.strip()
                    if not t.isdigit():
                        result.append(t)
                    elif i == 0:
                        result.append(t.zfill(2))
                    elif len(t) == 1:
                        result.append(t + "0")
                    else:
                        result.append(t.zfill(2))
                return ".".join(result)
            if isinstance(val, int):
                return str(val).zfill(2)
            if isinstance(val, float):
                if math.isnan(val) or math.isinf(val):
                    return ""
                s = str(val)
                if "." in s:
                    integer_part, decimal_part = s.split(".", 1)
                    if decimal_part == "0":
                        return integer_part.zfill(2)
                    if len(decimal_part) == 1:
                        decimal_part = decimal_part + "0"
                    return integer_part.zfill(2) + "." + decimal_part.zfill(2)
                return s.zfill(2)
            return normalizar_codigo(str(val))

        def parse_float(val) -> float:
            if val is None:
                return 0.0
            try:
                return float(str(val).replace(",", "."))
            except:
                return 0.0

        def parse_fecha(val) -> Optional[datetime]:
            if not val:
                return None
            if isinstance(val, datetime):
                return val
            try:
                return datetime.strptime(str(val)[:10], "%Y-%m-%d")
            except:
                return None

        def parse_fecha_formula(formula: str, fecha_inicio_base) -> Optional[datetime]:
            import re
            from datetime import timedelta
            if not formula or not isinstance(formula, str) or not formula.startswith("="):
                return None
            if fecha_inicio_base is None:
                return None
            m = re.match(r'^=\+?[A-Za-z]+\d+\+(\d+)(?:-(\d+))?$', formula.strip())
            if m:
                add_days = int(m.group(1))
                sub_days = int(m.group(2)) if m.group(2) else 0
                return fecha_inicio_base + timedelta(days=add_days - sub_days)
            return None

        # ── Procesar filas ──────────────────────────────────────────────────────

        valid_partidas = 0
        invalid_rows = 0
        total_rows = 0

        if is_old_xls:
            # ── Formato SGP (.xls) ──────────────────────────────────────────
            # Columnas A-K obligatorias:
            #   A=comisaria B=proveedor C=codigo D=partida E=und
            #   F=metrado G=pu H=parcial I=met_eje J=inicio K=fec_termino
            import xlrd as xlrd_lib
            wb = xlrd_lib.open_workbook(file_contents=contents)
            ws = wb.sheet_by_index(0)
            datemode = wb.datemode

            if ws.ncols < 11:
                raise HTTPException(
                    status_code=400,
                    detail=f"El archivo debe tener al menos 11 columnas (A-K). Solo tiene {ws.ncols}."
                )

            header = [str(ws.cell_value(0, j)).strip().lower() for j in range(min(11, ws.ncols))]
            print(f"📋 Encabezados SGP detectados: {header}")

            def parse_fecha_sgp(val) -> Optional[datetime]:
                """Convierte serial de Excel o string a datetime."""
                if not val and val != 0:
                    return None
                if isinstance(val, (int, float)) and val > 1000:
                    try:
                        return xlrd_lib.xldate_as_datetime(float(val), datemode)
                    except:
                        return None
                if isinstance(val, str) and val.strip():
                    try:
                        return datetime.strptime(val.strip()[:10], "%Y-%m-%d")
                    except:
                        return None
                return None

            for i in range(1, ws.nrows):
                total_rows += 1
                row = [ws.cell_value(i, j) for j in range(ws.ncols)]

                if all(not v for v in row):
                    invalid_rows += 1
                    continue

                try:
                    # C(2)=codigo, D(3)=partida, E(4)=und, F(5)=metrado,
                    # G(6)=pu, H(7)=parcial, J(9)=inicio, K(10)=fec_termino
                    codigo_partida  = normalizar_codigo(row[2] if len(row) > 2 else None)
                    descripcion     = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    unidad          = str(row[4]).strip() if len(row) > 4 and row[4] else "UND"
                    metrado         = parse_float(row[5] if len(row) > 5 else None)
                    precio_unitario = parse_float(row[6] if len(row) > 6 else None)
                    precio_total    = parse_float(row[7] if len(row) > 7 else None)
                    if precio_total == 0.0 and metrado > 0 and precio_unitario > 0:
                        precio_total = round(metrado * precio_unitario, 2)
                    fecha_inicio    = parse_fecha_sgp(row[9]  if len(row) > 9  else None)
                    fecha_fin       = parse_fecha_sgp(row[10] if len(row) > 10 else None)

                    if codigo_partida and descripcion:
                        db.add(PartidaModel(
                            cronograma_id=cronograma_id,
                            codigo_interno=codigo_partida,
                            comisaria_id=comisaria_id,
                            codigo_partida=codigo_partida,
                            descripcion=descripcion,
                            unidad=unidad,
                            metrado=metrado,
                            precio_unitario=precio_unitario,
                            precio_total=precio_total,
                            fecha_inicio=fecha_inicio,
                            fecha_fin=fecha_fin,
                            nivel_jerarquia=calcular_nivel_jerarquia(codigo_partida),
                            partida_padre=get_partida_padre(codigo_partida),
                            created_at=datetime.now()
                        ))
                        valid_partidas += 1
                    else:
                        invalid_rows += 1

                except Exception as e:
                    print(f"Error fila {total_rows}: {e}")
                    invalid_rows += 1

        else:
            # ── Formatos anteriores (.xlsx) ─────────────────────────────────
            workbook      = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            workbook_fmls = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
            sheet      = workbook.active
            sheet_fmls = workbook_fmls.active

            header_row = [str(c).strip().upper() if c else "" for c in
                          next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            print(f"📋 Encabezados detectados: {header_row}")

            header_lower = [h.lower() for h in header_row]
            # SGP dentro de xlsx (misma estructura A-K)
            if len(header_lower) >= 11 and header_lower[0] == "comisaria" and header_lower[2] == "codigo":
                fmt = "SGP_XLSX"
            elif header_row and header_row[0] in ("ITEM", "N°", "NRO", "NUM", "#"):
                fmt = "A" if "DESCRIPCION" in header_row[:3] else "B"
            else:
                fmt = "B"
            print(f"📊 Formato detectado: {fmt}")

            rows_vals = list(sheet.iter_rows(min_row=2, values_only=True))
            rows_fmls = list(sheet_fmls.iter_rows(min_row=2, values_only=True))

            for idx, row in enumerate(rows_vals):
                total_rows += 1
                if not row or all(c is None for c in row):
                    invalid_rows += 1
                    continue

                fml_row = rows_fmls[idx] if idx < len(rows_fmls) else ()

                try:
                    if fmt == "SGP_XLSX":
                        codigo_partida  = normalizar_codigo(row[2] if len(row) > 2 else None)
                        descripcion     = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        unidad          = str(row[4]).strip() if len(row) > 4 and row[4] else "UND"
                        metrado         = parse_float(row[5] if len(row) > 5 else None)
                        precio_unitario = parse_float(row[6] if len(row) > 6 else None)
                        precio_total    = parse_float(row[7] if len(row) > 7 else None)
                        if precio_total == 0.0 and metrado > 0 and precio_unitario > 0:
                            precio_total = round(metrado * precio_unitario, 2)
                        fecha_inicio    = parse_fecha(row[9]  if len(row) > 9  else None)
                        fecha_fin       = parse_fecha(row[10] if len(row) > 10 else None)
                        codigo_interno  = codigo_partida
                    elif fmt == "A":
                        codigo_partida  = normalizar_codigo(row[0])
                        descripcion     = str(row[1]).strip() if row[1] else ""
                        unidad          = str(row[2]).strip() if row[2] else "UND"
                        metrado         = parse_float(row[3])
                        precio_unitario = parse_float(row[4])
                        precio_total    = parse_float(row[5])
                        if precio_total == 0.0 and metrado > 0 and precio_unitario > 0:
                            precio_total = round(metrado * precio_unitario, 2)
                        fecha_inicio    = parse_fecha(row[6] if len(row) > 6 else None)
                        fecha_fin       = parse_fecha(row[7] if len(row) > 7 else None)
                        if fecha_fin is None and len(fml_row) > 7:
                            fecha_fin = parse_fecha_formula(fml_row[7], fecha_inicio)
                        codigo_interno  = codigo_partida
                    else:  # fmt == "B"
                        codigo_interno  = str(row[1]).strip() if row[1] else f"AUTO_{total_rows}"
                        codigo_partida  = normalizar_codigo(row[3])
                        descripcion     = str(row[4]).strip() if row[4] else ""
                        metrado         = parse_float(row[6]) if len(row) > 6 else 0.0
                        precio_unitario = parse_float(row[7]) if len(row) > 7 else 0.0
                        precio_total    = parse_float(row[8]) if len(row) > 8 else 0.0
                        unidad          = str(row[9]).strip() if len(row) > 9 and row[9] else "UND"
                        fecha_inicio    = parse_fecha(row[10] if len(row) > 10 else None)
                        fecha_fin       = parse_fecha(row[11] if len(row) > 11 else None)
                        if fecha_fin is None and len(fml_row) > 11:
                            fecha_fin = parse_fecha_formula(fml_row[11], fecha_inicio)

                    if codigo_partida and descripcion:
                        db.add(PartidaModel(
                            cronograma_id=cronograma_id,
                            codigo_interno=codigo_interno,
                            comisaria_id=comisaria_id,
                            codigo_partida=codigo_partida,
                            descripcion=descripcion,
                            unidad=unidad,
                            metrado=metrado,
                            precio_unitario=precio_unitario,
                            precio_total=precio_total,
                            fecha_inicio=fecha_inicio,
                            fecha_fin=fecha_fin,
                            nivel_jerarquia=calcular_nivel_jerarquia(codigo_partida),
                            partida_padre=get_partida_padre(codigo_partida),
                            created_at=datetime.now()
                        ))
                        valid_partidas += 1
                    else:
                        invalid_rows += 1

                except Exception as e:
                    print(f"Error procesando fila {total_rows}: {e}")
                    invalid_rows += 1

        await db.commit()

        # ── Rollup: precio_total, fecha_inicio y fecha_fin de agrupaciones ──────
        stmt_all = select(PartidaModel).where(PartidaModel.cronograma_id == cronograma_id)
        result_all = await db.execute(stmt_all)
        todas = list(result_all.scalars().all())

        for partida in sorted(todas, key=lambda p: p.nivel_jerarquia, reverse=True):
            hijos = [p for p in todas if p.partida_padre == partida.codigo_partida]
            if not hijos:
                continue
            total_hijos = sum(h.precio_total or 0.0 for h in hijos)
            if total_hijos > 0:
                partida.precio_total = total_hijos
            fechas_ini = [h.fecha_inicio for h in hijos if h.fecha_inicio]
            fechas_fin = [h.fecha_fin    for h in hijos if h.fecha_fin]
            if fechas_ini:
                partida.fecha_inicio = min(fechas_ini)
            if fechas_fin:
                partida.fecha_fin = max(fechas_fin)

        await db.commit()

        print(f"✅ Cronograma importado: {valid_partidas} partidas válidas de {total_rows} filas")

        return ImportStats(
            total_rows=total_rows,
            valid_partidas=valid_partidas,
            invalid_rows=invalid_rows,
            cronograma_id=cronograma_id,
            stats={
                "nombre_cronograma": nombre_cronograma,
                "comisaria_id": comisaria_id,
                "archivo": file.filename
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback, sys
        error_type = type(e).__name__
        tb = traceback.format_exc()
        print(f"❌ Error: {error_type}: {e}")
        print(tb)
        sys.stdout.flush()
        raise HTTPException(status_code=500, detail=f"Error [{error_type}]: {str(e)}")

@router.delete("/{cronograma_id}", status_code=204)
async def delete_cronograma(cronograma_id: int, db: AsyncSession = Depends(get_db)):
    """
    Eliminar cronograma y todas sus partidas

    Args:
        cronograma_id: ID del cronograma a eliminar
    """
    # Verificar que existe
    stmt = select(CronogramaModel).where(CronogramaModel.id == cronograma_id)
    result = await db.execute(stmt)
    cronograma = result.scalar_one_or_none()

    if not cronograma:
        raise HTTPException(status_code=404, detail="Cronograma no encontrado")

    cronograma_name = cronograma.nombre

    # Eliminar todas las partidas del cronograma
    stmt = delete(PartidaModel).where(PartidaModel.cronograma_id == cronograma_id)
    await db.execute(stmt)

    # Eliminar el cronograma
    stmt = delete(CronogramaModel).where(CronogramaModel.id == cronograma_id)
    await db.execute(stmt)

    await db.commit()

    print(f"✅ Cronograma eliminado: {cronograma_name} (ID: {cronograma_id})")

@router.get("/{cronograma_id}/partidas", response_model=List[PartidaResponse])
async def get_cronograma_partidas(cronograma_id: int, db: AsyncSession = Depends(get_db)):
    """
    Obtener solo las partidas de un cronograma específico

    Args:
        cronograma_id: ID del cronograma

    Returns:
        List[PartidaResponse]: Lista de partidas del cronograma
    """
    print(f"🔍 DEBUG: GET /cronogramas/{cronograma_id}/partidas")

    # Verificar que el cronograma existe
    stmt_cronograma = select(CronogramaModel).where(CronogramaModel.id == cronograma_id)
    result_cronograma = await db.execute(stmt_cronograma)
    cronograma = result_cronograma.scalar_one_or_none()

    if not cronograma:
        raise HTTPException(status_code=404, detail="Cronograma no encontrado")

    # Obtener partidas del cronograma
    stmt = select(PartidaModel).where(
        PartidaModel.cronograma_id == cronograma_id
    ).order_by(PartidaModel.codigo_partida)

    result = await db.execute(stmt)
    partidas = result.scalars().all()

    print(f"📊 Encontradas {len(partidas)} partidas para cronograma {cronograma_id}")

    return [partida_model_to_response(partida) for partida in partidas]

@router.get("/comisaria/{comisaria_id}/detalle", response_model=CronogramaWithPartidas)
async def get_cronograma_detalle_by_comisaria(comisaria_id: int, db: AsyncSession = Depends(get_db)):
    """
    Obtener cronograma más reciente de una comisaría CON sus partidas

    Args:
        comisaria_id: ID de la comisaría

    Returns:
        CronogramaWithPartidas: Cronograma completo con todas las partidas
    """
    print(f"🔍 DEBUG: GET /cronogramas/comisaria/{comisaria_id}/detalle")

    # Obtener el cronograma más reciente de la comisaría
    stmt = select(CronogramaModel).where(
        CronogramaModel.comisaria_id == comisaria_id
    ).order_by(CronogramaModel.created_at.desc()).limit(1)

    result = await db.execute(stmt)
    cronograma = result.scalars().first()

    if not cronograma:
        raise HTTPException(status_code=404, detail="No se encontró cronograma para esta comisaría")

    # Obtener todas las partidas del cronograma
    stmt_partidas = select(PartidaModel).where(
        PartidaModel.cronograma_id == cronograma.id
    ).order_by(PartidaModel.codigo_partida)

    result_partidas = await db.execute(stmt_partidas)
    partidas = result_partidas.scalars().all()

    print(f"📊 Encontrado cronograma {cronograma.nombre} con {len(partidas)} partidas")

    # Convertir a modelos de respuesta
    cronograma_response = cronograma_model_to_response(cronograma)
    partidas_response = [partida_model_to_response(partida) for partida in partidas]

    return CronogramaWithPartidas(
        **cronograma_response.model_dump(),
        partidas=partidas_response
    )

@router.put("/partidas/{partida_id}/fechas", response_model=PartidaResponse)
async def update_partida_fechas(
    partida_id: int,
    fechas_data: PartidaFechasUpdate,
    db: AsyncSession = Depends(get_db)
):
    """
    Actualizar fechas de una partida específica

    Args:
        partida_id: ID de la partida
        fechas_data: Nuevas fechas de inicio y fin
        db: Sesión de base de datos

    Returns:
        PartidaResponse: Partida actualizada
    """
    print(f"🔗 PUT /cronogramas/partidas/{partida_id}/fechas")

    # Buscar la partida
    stmt = select(PartidaModel).where(PartidaModel.id == partida_id)
    result = await db.execute(stmt)
    partida = result.scalar_one_or_none()

    if not partida:
        raise HTTPException(status_code=404, detail="Partida no encontrada")

    # Actualizar fechas
    if fechas_data.fecha_inicio is not None:
        if fechas_data.fecha_inicio.strip():
            # Convertir a datetime sin timezone para mantener consistencia con BD
            fecha_dt = datetime.fromisoformat(fechas_data.fecha_inicio.replace('Z', '+00:00'))
            partida.fecha_inicio = fecha_dt.replace(tzinfo=None)
        else:
            partida.fecha_inicio = None

    if fechas_data.fecha_fin is not None:
        if fechas_data.fecha_fin.strip():
            # Convertir a datetime sin timezone para mantener consistencia con BD
            fecha_dt = datetime.fromisoformat(fechas_data.fecha_fin.replace('Z', '+00:00'))
            partida.fecha_fin = fecha_dt.replace(tzinfo=None)
        else:
            partida.fecha_fin = None

    await db.commit()

    # Recalcular fechas del cronograma padre si es necesario
    cronograma_id = partida.cronograma_id

    # Obtener todas las partidas del cronograma para recalcular fechas
    stmt_cronograma = select(CronogramaModel).where(CronogramaModel.id == cronograma_id)
    result_cronograma = await db.execute(stmt_cronograma)
    cronograma = result_cronograma.scalar_one_or_none()

    if cronograma:
        # Obtener todas las partidas del cronograma
        stmt_partidas = select(PartidaModel).where(PartidaModel.cronograma_id == cronograma_id)
        result_partidas = await db.execute(stmt_partidas)
        todas_partidas = result_partidas.scalars().all()

        # Recalcular fechas del cronograma
        fechas_inicio_validas = [p.fecha_inicio for p in todas_partidas if p.fecha_inicio]
        fechas_fin_validas = [p.fecha_fin for p in todas_partidas if p.fecha_fin]

        if fechas_inicio_validas:
            cronograma.fecha_inicio = min(fechas_inicio_validas)
        if fechas_fin_validas:
            cronograma.fecha_fin = max(fechas_fin_validas)

        await db.commit()

    print(f"✅ Fechas actualizadas para partida {partida_id}: inicio={fechas_data.fecha_inicio}, fin={fechas_data.fecha_fin}")

    return partida_model_to_response(partida)