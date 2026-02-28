"""
📊 CRONOGRAMAS API ROUTER - NEMAEC ERP
Endpoints para gestión de cronogramas valorizados con importación Excel.
"""
import json
import os
import io
from typing import List, Optional, Dict, Any
from datetime import datetime
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
import openpyxl

router = APIRouter(
    prefix="/cronogramas",
    tags=["cronogramas"],
    responses={404: {"description": "Not found"}}
)

# Archivo de persistencia
CRONOGRAMAS_FILE = "/tmp/nemaec_cronogramas.json"

# Modelos Pydantic

class PartidaBase(BaseModel):
    codigo_interno: str
    comisaria_id: int
    codigo_partida: str
    descripcion: str
    unidad: str
    metrado: float
    precio_unitario: float
    precio_total: float
    fecha_inicio: Optional[str] = None  # 🎯 PERMITIR FECHAS NULAS
    fecha_fin: Optional[str] = None     # 🎯 PERMITIR FECHAS NULAS
    nivel_jerarquia: int
    partida_padre: Optional[str] = None

class PartidaResponse(PartidaBase):
    id: int
    created_at: str

class CronogramaCreate(BaseModel):
    comisaria_id: int
    nombre_cronograma: Optional[str] = None

class CronogramaResponse(BaseModel):
    id: int
    comisaria_id: int
    nombre_cronograma: str
    archivo_original: str
    fecha_importacion: str
    total_presupuesto: float
    total_partidas: int
    fecha_inicio_obra: Optional[str] = None  # 🎯 PERMITIR FECHAS NULAS
    fecha_fin_obra: Optional[str] = None     # 🎯 PERMITIR FECHAS NULAS
    estado: str
    partidas: List[PartidaResponse]
    created_at: str

class CronogramaResumen(BaseModel):
    id: int
    comisaria_id: int
    comisaria_nombre: str
    nombre_cronograma: str
    total_presupuesto: float
    total_partidas: int
    fecha_inicio_obra: Optional[str] = None  # 🎯 PERMITIR FECHAS NULAS
    fecha_fin_obra: Optional[str] = None     # 🎯 PERMITIR FECHAS NULAS
    estado: str
    fecha_importacion: str

class ValidationError(BaseModel):
    errors: List[str]
    warnings: List[str]

class ExcelValidationResult(BaseModel):
    is_valid: bool
    errors: List[str]
    warnings: List[str]
    preview: List[PartidaBase]
    stats: Dict[str, Any]

# Funciones de persistencia

def load_cronogramas() -> List[Dict]:
    """Cargar cronogramas del archivo JSON"""
    if os.path.exists(CRONOGRAMAS_FILE):
        try:
            with open(CRONOGRAMAS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading cronogramas: {e}")
    return []

def save_cronogramas(cronogramas: List[Dict]):
    """Guardar cronogramas al archivo JSON"""
    try:
        with open(CRONOGRAMAS_FILE, 'w', encoding='utf-8') as f:
            json.dump(cronogramas, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving cronogramas: {e}")

def get_partida_padre(codigo_partida: str) -> Optional[str]:
    """Obtener código de partida padre"""
    partes = codigo_partida.split('.')
    if len(partes) <= 1:
        return None
    return '.'.join(partes[:-1])

# Variables globales
cronogramas_data = load_cronogramas()
next_cronograma_id = max([c["id"] for c in cronogramas_data], default=0) + 1
next_partida_id = 1

# Calcular next_partida_id basado en partidas existentes
for cronograma in cronogramas_data:
    for partida in cronograma.get("partidas", []):
        if partida.get("id", 0) >= next_partida_id:
            next_partida_id = partida["id"] + 1

# Endpoints

@router.get("/", response_model=List[CronogramaResumen])
async def get_all_cronogramas():
    """
    Obtener todos los cronogramas (resumen)

    Returns:
        List[CronogramaResumen]: Lista de cronogramas
    """
    resumenes = []
    for cronograma in cronogramas_data:
        resumenes.append({
            "id": cronograma["id"],
            "comisaria_id": cronograma["comisaria_id"],
            "comisaria_nombre": f"Comisaría {cronograma['comisaria_id']}",  # En producción vendría de join
            "nombre_cronograma": cronograma["nombre_cronograma"],
            "total_presupuesto": cronograma["total_presupuesto"],
            "total_partidas": cronograma["total_partidas"],
            "fecha_inicio_obra": cronograma["fecha_inicio_obra"],
            "fecha_fin_obra": cronograma["fecha_fin_obra"],
            "estado": cronograma["estado"],
            "fecha_importacion": cronograma["fecha_importacion"]
        })

    return sorted(resumenes, key=lambda x: x["fecha_importacion"], reverse=True)

@router.get("/comisaria/{comisaria_id}", response_model=CronogramaResponse)
async def get_cronograma_by_comisaria(comisaria_id: int):
    """
    Obtener cronograma por comisaría

    Args:
        comisaria_id: ID de la comisaría

    Returns:
        CronogramaResponse: Cronograma completo
    """
    # Obtener el cronograma más reciente para esta comisaría
    cronogramas_comisaria = [c for c in cronogramas_data if c["comisaria_id"] == comisaria_id]
    if not cronogramas_comisaria:
        cronograma = None
    else:
        # Ordenar por fecha de importación (más reciente primero)
        cronograma = sorted(cronogramas_comisaria, key=lambda x: x["fecha_importacion"], reverse=True)[0]

    if not cronograma:
        raise HTTPException(status_code=404, detail="Cronograma no encontrado")

    return cronograma

@router.post("/validate-excel", response_model=ExcelValidationResult)
async def validate_excel_file(file: UploadFile = File(...)):
    """
    Validar archivo Excel antes de importación

    Args:
        file: Archivo Excel a validar

    Returns:
        ExcelValidationResult: Resultado de validación
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser de formato Excel (.xlsx o .xls)"
        )

    try:
        # Leer archivo Excel
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active

        errors = []
        warnings = []
        valid_partidas = []
        total_presupuesto = 0

        # Primero, leer los headers para debug
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"DEBUG - Headers encontrados: {header_row}")

        # Procesar filas (asumiendo que la primera fila son headers)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Debug: mostrar estructura de las primeras 3 filas
                if row_num <= 4:
                    print(f"DEBUG - Fila {row_num}: {row}")

                # Mapeo de columnas basado en la estructura del Excel
                if len(row) < 5:
                    errors.append(f"Fila {row_num}: Estructura de fila incompleta (solo {len(row)} columnas)")
                    continue

                # Intentar mapear columnas de manera más flexible
                codigo_interno = None
                codigo_partida = None
                descripcion = None
                metrado = 0
                precio_unitario = 0
                precio_total = 0
                unidad = "UND"

                # Buscar valores no vacíos en las primeras columnas
                for i, cell in enumerate(row[:15]):  # Revisar hasta 15 columnas
                    if cell is not None and str(cell).strip():
                        if i == 0 and not codigo_interno:
                            codigo_interno = cell
                        elif i == 1 and not codigo_partida:
                            codigo_partida = cell
                        elif i == 2 and not descripcion:
                            descripcion = cell
                        elif isinstance(cell, (int, float)) and cell > 0:
                            if metrado == 0:
                                metrado = cell
                            elif precio_unitario == 0:
                                precio_unitario = cell
                            elif precio_total == 0:
                                precio_total = cell

                # Validaciones básicas
                if not codigo_interno:
                    errors.append(f"Fila {row_num}: Código interno es obligatorio")
                    if row_num <= 4:  # Debug solo para las primeras filas
                        print(f"DEBUG - Fila {row_num} sin código interno. Row: {row[:5]}")
                    continue

                if not codigo_partida:
                    errors.append(f"Fila {row_num}: Código de partida es obligatorio")
                    if row_num <= 4:
                        print(f"DEBUG - Fila {row_num} sin código partida. Row: {row[:5]}")
                    continue

                if not descripcion:
                    errors.append(f"Fila {row_num}: Descripción es obligatoria")
                    if row_num <= 4:
                        print(f"DEBUG - Fila {row_num} sin descripción. Row: {row[:5]}")
                    continue

                # Validar precio_total
                if precio_total <= 0:
                    errors.append(f"Fila {row_num}: Precio total inválido ({precio_total})")
                    if row_num <= 4:
                        print(f"DEBUG - Fila {row_num} precio total inválido. Row: {row[:10]}")
                    continue

                # Intentar obtener fechas del Excel si están disponibles
                fecha_inicio = "2026-01-01T00:00:00Z"  # Valor por defecto
                fecha_fin = "2026-12-31T00:00:00Z"      # Valor por defecto

                if len(row) > 11:
                    try:
                        fecha_inicio_excel = row[10]  # Columna K
                        fecha_fin_excel = row[11]     # Columna L

                        if fecha_inicio_excel and isinstance(fecha_inicio_excel, datetime):
                            fecha_inicio = fecha_inicio_excel.isoformat() + "Z"

                        if fecha_fin_excel and isinstance(fecha_fin_excel, datetime):
                            fecha_fin = fecha_fin_excel.isoformat() + "Z"
                    except:
                        pass  # Mantener valores por defecto si hay error

                # Crear partida válida
                partida = PartidaBase(
                    codigo_interno=str(codigo_interno),
                    comisaria_id=0,  # Se asignará al importar
                    codigo_partida=str(codigo_partida),
                    descripcion=str(descripcion),
                    unidad=str(unidad),
                    metrado=float(metrado or 1),
                    precio_unitario=float(precio_unitario or precio_total),
                    precio_total=float(precio_total),
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    nivel_jerarquia=len(str(codigo_partida).split('.')),
                    partida_padre=get_partida_padre(str(codigo_partida))
                )

                valid_partidas.append(partida)
                total_presupuesto += partida.precio_total

            except Exception as e:
                errors.append(f"Fila {row_num}: Error procesando fila - {str(e)}")

        return ExcelValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            warnings=warnings,
            preview=valid_partidas[:10],  # Solo primeras 10 para preview
            stats={
                "total_filas": sheet.max_row - 1,  # Excluyendo header
                "partidas_validas": len(valid_partidas),
                "total_presupuesto": total_presupuesto
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error al procesar archivo Excel: {str(e)}"
        )

@router.post("/import", response_model=CronogramaResponse)
async def import_cronograma(
    comisaria_id: int = Form(...),
    nombre_cronograma: Optional[str] = Form(None),
    file: UploadFile = File(...)
):
    """
    Importar cronograma desde Excel

    Args:
        comisaria_id: ID de la comisaría
        nombre_cronograma: Nombre del cronograma (opcional)
        file: Archivo Excel

    Returns:
        CronogramaResponse: Cronograma importado
    """
    global next_cronograma_id, next_partida_id

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser de formato Excel (.xlsx o .xls)"
        )

    try:
        # Leer y procesar archivo Excel
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active

        partidas = []
        total_presupuesto = 0
        fechas_inicio = []
        fechas_fin = []

        # Detectar si el cronograma existente tiene fechas para mantener
        cronograma_existente = next(
            (c for c in cronogramas_data if c["comisaria_id"] == comisaria_id),
            None
        )

        # Procesar todas las filas
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if len(row) < 10:
                    continue

                codigo_interno = row[1]
                codigo_partida = row[3]
                descripcion = row[4]
                metrado = row[6] or 0
                precio_unitario = row[7] or 0
                precio_total = row[8] or 0
                unidad = row[9] or "UND"

                if not all([codigo_interno, codigo_partida, descripcion]):
                    continue

                # 🎯 LÓGICA INTELIGENTE DE FECHAS
                fecha_inicio = None
                fecha_fin = None

                # 1. Intentar obtener fechas del Excel (columnas K=10, L=11)
                if len(row) > 11:
                    try:
                        fecha_inicio_excel = row[10]  # Columna K (fecha inicio)
                        fecha_fin_excel = row[11]     # Columna L (fecha fin)

                        # Validar si son fechas válidas
                        if fecha_inicio_excel and isinstance(fecha_inicio_excel, datetime):
                            fecha_inicio = fecha_inicio_excel.isoformat() + "Z"
                        elif fecha_inicio_excel and str(fecha_inicio_excel).strip():
                            # Intentar parsear como texto
                            try:
                                fecha_parseada = datetime.strptime(str(fecha_inicio_excel).strip(), "%Y-%m-%d")
                                fecha_inicio = fecha_parseada.isoformat() + "Z"
                            except:
                                pass

                        if fecha_fin_excel and isinstance(fecha_fin_excel, datetime):
                            fecha_fin = fecha_fin_excel.isoformat() + "Z"
                        elif fecha_fin_excel and str(fecha_fin_excel).strip():
                            # Intentar parsear como texto
                            try:
                                fecha_parseada = datetime.strptime(str(fecha_fin_excel).strip(), "%Y-%m-%d")
                                fecha_fin = fecha_parseada.isoformat() + "Z"
                            except:
                                pass
                    except Exception as e:
                        print(f"DEBUG - Error procesando fechas en fila {row_num}: {e}")

                # Debug: mostrar fechas extraídas del Excel
                if row_num <= 5 and (fecha_inicio or fecha_fin):
                    print(f"DEBUG - Fila {row_num} - Fechas del Excel: inicio={fecha_inicio}, fin={fecha_fin}")

                # 2. Solo buscar en cronograma existente si NO hay fechas en el Excel
                if (not fecha_inicio or not fecha_fin) and cronograma_existente:
                    # Buscar partida existente con mismo código
                    partida_existente = next(
                        (p for p in cronograma_existente.get("partidas", [])
                         if p.get("codigo_partida") == str(codigo_partida)),
                        None
                    )
                    if partida_existente:
                        if not fecha_inicio:
                            fecha_inicio = partida_existente.get("fecha_inicio")
                        if not fecha_fin:
                            fecha_fin = partida_existente.get("fecha_fin")

                # 3. Las fechas del Excel ya están asignadas arriba
                # Solo necesitamos manejar None si no hay fechas disponibles
                # (no necesitamos reasignar None aquí)

                # Crear partida
                partida = {
                    "id": next_partida_id,
                    "codigo_interno": str(codigo_interno),
                    "comisaria_id": comisaria_id,
                    "codigo_partida": str(codigo_partida),
                    "descripcion": str(descripcion),
                    "unidad": str(unidad),
                    "metrado": float(metrado),
                    "precio_unitario": float(precio_unitario),
                    "precio_total": float(precio_total),
                    "fecha_inicio": fecha_inicio,
                    "fecha_fin": fecha_fin,
                    "nivel_jerarquia": len(str(codigo_partida).split('.')),
                    "partida_padre": get_partida_padre(str(codigo_partida)),
                    "created_at": datetime.now().isoformat()
                }

                partidas.append(partida)
                total_presupuesto += partida["precio_total"]
                next_partida_id += 1

            except Exception as e:
                print(f"Error procesando fila {row_num}: {e}")
                continue

        if not partidas:
            raise HTTPException(
                status_code=400,
                detail="No se pudieron procesar partidas del archivo Excel"
            )

        # Calcular fechas de obra dinámicamente
        fechas_inicio_validas = [p["fecha_inicio"] for p in partidas if p.get("fecha_inicio")]
        fechas_fin_validas = [p["fecha_fin"] for p in partidas if p.get("fecha_fin")]

        fecha_inicio_obra = None
        fecha_fin_obra = None

        if fechas_inicio_validas:
            fecha_inicio_obra = min(fechas_inicio_validas)
        if fechas_fin_validas:
            fecha_fin_obra = max(fechas_fin_validas)

        # Crear cronograma
        cronograma = {
            "id": next_cronograma_id,
            "comisaria_id": comisaria_id,
            "nombre_cronograma": nombre_cronograma or f"Cronograma {file.filename.replace('.xlsx', '')}",
            "archivo_original": file.filename,
            "fecha_importacion": datetime.now().isoformat(),
            "total_presupuesto": total_presupuesto,
            "total_partidas": len(partidas),
            "fecha_inicio_obra": fecha_inicio_obra,  # Calculado dinámicamente
            "fecha_fin_obra": fecha_fin_obra,        # Calculado dinámicamente
            "estado": "activo",
            "partidas": partidas,
            "created_at": datetime.now().isoformat()
        }

        cronogramas_data.append(cronograma)
        save_cronogramas(cronogramas_data)
        next_cronograma_id += 1

        print(f"✅ Cronograma importado: {cronograma['nombre_cronograma']} (ID: {cronograma['id']})")
        print(f"   Partidas: {len(partidas)}, Presupuesto: S/ {total_presupuesto:,.2f}")

        return cronograma

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al importar cronograma: {str(e)}"
        )

@router.delete("/{cronograma_id}", status_code=204)
async def delete_cronograma(cronograma_id: int):
    """
    Eliminar cronograma

    Args:
        cronograma_id: ID del cronograma
    """
    cronograma_index = next(
        (i for i, c in enumerate(cronogramas_data) if c["id"] == cronograma_id),
        None
    )

    if cronograma_index is None:
        raise HTTPException(status_code=404, detail="Cronograma no encontrado")

    removed = cronogramas_data.pop(cronograma_index)
    save_cronogramas(cronogramas_data)

    print(f"✅ Cronograma eliminado: {removed['nombre_cronograma']} (ID: {cronograma_id})")